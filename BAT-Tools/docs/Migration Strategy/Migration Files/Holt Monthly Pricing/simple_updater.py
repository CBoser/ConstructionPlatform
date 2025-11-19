#!/usr/bin/env python3
"""
SIMPLE UNIVERSAL BAT UPDATER
A simplified version that works with any BAT file structure
"""

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import sys


def update_bat_file(file_path):
    """Simple updater for any BAT file"""
    print("\n" + "="*60)
    print("SIMPLE BAT UPDATER")
    print("="*60)
    print(f"File: {os.path.basename(file_path)}")
    print("-"*60)
    
    # Load file
    print("Loading file...")
    try:
        if file_path.endswith('.xlsm'):
            wb = openpyxl.load_workbook(file_path, keep_vba=True)
        else:
            wb = openpyxl.load_workbook(file_path)
        print("✔ File loaded")
    except Exception as e:
        print(f"❌ Error: {e}")
        return False
    
    # Find sheets
    margin_sheet = None
    pricing_sheet = None
    
    # Look for margin updates sheet
    for sheet_name in wb.sheetnames:
        lower_name = sheet_name.lower()
        if 'margin' in lower_name or 'update' in lower_name:
            margin_sheet = wb[sheet_name]
            print(f"Found margin sheet: {sheet_name}")
            break
    
    # Look for pricing sheet
    for sheet_name in wb.sheetnames:
        lower_name = sheet_name.lower()
        if 'pricing' in lower_name or 'cost' in lower_name:
            pricing_sheet = wb[sheet_name]
            print(f"Found pricing sheet: {sheet_name}")
            break
    
    if not margin_sheet or not pricing_sheet:
        print("❌ Required sheets not found")
        return False
    
    # Read updates (simplified)
    print("\nReading updates...")
    updates = []
    
    # Start from row 5 (common start row)
    for row in range(5, margin_sheet.max_row + 1):
        zone = margin_sheet.cell(row=row, column=1).value
        category = margin_sheet.cell(row=row, column=2).value
        margin = margin_sheet.cell(row=row, column=6).value
        
        if zone and category and margin is not None:
            updates.append({
                'zone': str(zone).strip(),
                'category': str(category).strip(),
                'margin': float(margin) if margin <= 1 else float(margin)/100,
                'row': row
            })
    
    print(f"Found {len(updates)} updates")
    
    if not updates:
        print("No updates to apply")
        return False
    
    # Apply updates (simplified)
    print("\nApplying updates...")
    highlight = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    items_updated = 0
    
    for update in updates:
        matches = 0
        for row in range(2, pricing_sheet.max_row + 1):
            row_zone = str(pricing_sheet.cell(row=row, column=1).value or '').strip()
            row_category = str(pricing_sheet.cell(row=row, column=2).value or '').strip()
            
            if row_zone == update['zone'] and row_category == update['category']:
                # Update margin cells (columns 20, 23, 26, etc.)
                margin_cols = [20, 23, 26, 29, 32, 35, 38, 41, 44, 47, 50, 53]
                for col in margin_cols:
                    if col - 1 <= pricing_sheet.max_column:
                        cell = pricing_sheet.cell(row=row, column=col-1)
                        cell.value = update['margin']
                        cell.number_format = '0.0%'
                        cell.fill = highlight
                matches += 1
        
        items_updated += matches
        
        # Update status
        margin_sheet.cell(row=update['row'], column=9).value = f"✔ {matches} items"
    
    print(f"Updated {items_updated} items")
    
    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    base_name = os.path.splitext(file_path)[0]
    ext = os.path.splitext(file_path)[1]
    output_file = f"{base_name}_UPDATED_{timestamp}{ext}"
    
    print(f"\nSaving: {os.path.basename(output_file)}")
    try:
        wb.save(output_file)
        print("✔ File saved successfully")
        
        print("\n" + "="*60)
        print("UPDATE COMPLETE!")
        print(f"Items updated: {items_updated}")
        print("="*60)
        return True
    except Exception as e:
        print(f"❌ Error saving: {e}")
        return False


def main():
    if len(sys.argv) < 2:
        # Default file
        file_path = r"C:\Users\corey.boser\Documents\Holt Monthly Pricing\HOLT BAT OCTOBER 2025 9-29-25.xlsm"
        if not os.path.exists(file_path):
            print("Please provide file path:")
            print("python simple_updater.py 'path\\to\\file.xlsm'")
            sys.exit(1)
    else:
        file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        sys.exit(1)
    
    success = update_bat_file(file_path)
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
