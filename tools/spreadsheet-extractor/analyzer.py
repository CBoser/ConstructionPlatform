"""
Spreadsheet Analyzer Module
Extracts formulas, structure, and metadata from Excel files.
"""

import re
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Set
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


@dataclass
class CellInfo:
    address: str
    value: Any
    formula: Optional[str] = None
    data_type: str = "unknown"
    number_format: Optional[str] = None


@dataclass
class FormulaInfo:
    address: str
    sheet: str
    formula: str
    result: Any = None
    dependencies: List[str] = field(default_factory=list)
    functions: List[str] = field(default_factory=list)
    is_array_formula: bool = False


@dataclass
class SheetInfo:
    name: str
    used_range: str
    row_count: int
    column_count: int
    formula_count: int
    cell_count: int
    has_tables: bool = False


@dataclass
class NamedRangeInfo:
    name: str
    scope: str  # 'workbook' or sheet name
    refers_to: str
    is_formula: bool = False


@dataclass
class TableInfo:
    name: str
    sheet: str
    range: str
    headers: List[str] = field(default_factory=list)
    row_count: int = 0
    has_calculated_columns: bool = False


@dataclass
class AnalysisResult:
    file_name: str
    file_size: int
    sheets: List[SheetInfo] = field(default_factory=list)
    formulas: List[FormulaInfo] = field(default_factory=list)
    named_ranges: List[NamedRangeInfo] = field(default_factory=list)
    tables: List[TableInfo] = field(default_factory=list)
    function_stats: Dict[str, int] = field(default_factory=dict)
    complexity: str = "simple"
    errors: List[str] = field(default_factory=list)


# Excel functions that indicate dynamic arrays
ARRAY_FUNCTIONS = {
    'FILTER', 'SORT', 'SORTBY', 'UNIQUE', 'SEQUENCE', 'RANDARRAY',
    'XLOOKUP', 'XMATCH', 'LET', 'LAMBDA', 'MAP', 'REDUCE', 'SCAN',
    'MAKEARRAY', 'BYROW', 'BYCOL', 'CHOOSECOLS', 'CHOOSEROWS',
    'DROP', 'TAKE', 'EXPAND', 'WRAPCOLS', 'WRAPROWS', 'TOCOL', 'TOROW',
    'TEXTSPLIT', 'TEXTBEFORE', 'TEXTAFTER', 'VSTACK', 'HSTACK'
}


def extract_functions(formula: str) -> List[str]:
    """Extract Excel function names from a formula."""
    pattern = r'([A-Z_][A-Z0-9_]*)\s*\('
    matches = re.findall(pattern, formula.upper())
    return list(set(matches))


def extract_cell_references(formula: str) -> List[str]:
    """Extract cell references from a formula."""
    # Match cell refs like A1, $A$1, Sheet1!A1, 'Sheet Name'!A1
    pattern = r"(?:'[^']+!'?|[A-Za-z0-9_]+!)?\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?"
    matches = re.findall(pattern, formula.upper())
    return list(set(matches))


def is_array_formula(formula: str) -> bool:
    """Check if formula uses dynamic array functions."""
    upper_formula = formula.upper()
    return any(f'{func}(' in upper_formula for func in ARRAY_FUNCTIONS)


def calculate_complexity(result: AnalysisResult) -> str:
    """Calculate overall complexity score."""
    score = 0

    # Sheet count
    score += len(result.sheets) * 2

    # Formula count
    formula_count = len(result.formulas)
    if formula_count > 100:
        score += 30
    elif formula_count > 50:
        score += 20
    elif formula_count > 20:
        score += 10
    else:
        score += formula_count * 0.5

    # Unique functions
    score += len(result.function_stats) * 2

    # Array formulas
    array_count = sum(1 for f in result.formulas if f.is_array_formula)
    score += array_count * 5

    # Named ranges
    score += len(result.named_ranges) * 2

    # Tables
    score += len(result.tables) * 3

    if score > 100:
        return "very-complex"
    elif score > 50:
        return "complex"
    elif score > 20:
        return "moderate"
    return "simple"


class SpreadsheetAnalyzer:
    """Analyzes Excel spreadsheets to extract business logic."""

    def __init__(self):
        self.workbook = None
        self.file_path: Optional[Path] = None

    def analyze(self, file_path: str) -> AnalysisResult:
        """Analyze an Excel file and return structured results."""
        self.file_path = Path(file_path)

        if not self.file_path.exists():
            return AnalysisResult(
                file_name=self.file_path.name,
                file_size=0,
                errors=[f"File not found: {file_path}"]
            )

        file_size = self.file_path.stat().st_size
        suffix = self.file_path.suffix.lower()

        result = AnalysisResult(
            file_name=self.file_path.name,
            file_size=file_size
        )

        try:
            if suffix in ['.xlsx', '.xlsm', '.xlsb']:
                self._analyze_xlsx(result)
            elif suffix == '.xls':
                self._analyze_xls(result)
            elif suffix == '.csv':
                self._analyze_csv(result)
            else:
                result.errors.append(f"Unsupported file format: {suffix}")
        except Exception as e:
            result.errors.append(f"Error analyzing file: {str(e)}")

        # Calculate complexity
        result.complexity = calculate_complexity(result)

        return result

    def _analyze_xlsx(self, result: AnalysisResult):
        """Analyze .xlsx/.xlsm files using openpyxl."""
        if openpyxl is None:
            result.errors.append("openpyxl not installed. Run: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(
                self.file_path,
                data_only=False,  # Get formulas, not just values
                read_only=False
            )
        except Exception as e:
            result.errors.append(f"Failed to open workbook: {str(e)}")
            return

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            formula_count = 0
            cell_count = 0

            # Get dimensions
            if ws.dimensions:
                try:
                    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ws.dimensions)
                    row_count = max_row - min_row + 1
                    col_count = max_col - min_col + 1
                except:
                    row_count = ws.max_row or 0
                    col_count = ws.max_column or 0
            else:
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0

            # Scan cells for formulas
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_count += 1

                    # Check for formula
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and cell.value.startswith('=')):
                        formula_str = cell.value if isinstance(cell.value, str) else str(cell.value)
                        if formula_str.startswith('='):
                            formula_str = formula_str[1:]  # Remove leading =

                        formula_count += 1

                        # Extract formula info
                        functions = extract_functions(formula_str)
                        dependencies = extract_cell_references(formula_str)

                        formula_info = FormulaInfo(
                            address=cell.coordinate,
                            sheet=sheet_name,
                            formula=formula_str,
                            result=None,  # Would need data_only=True for this
                            dependencies=dependencies,
                            functions=functions,
                            is_array_formula=is_array_formula(formula_str)
                        )
                        result.formulas.append(formula_info)

                        # Update function stats
                        for func in functions:
                            result.function_stats[func] = result.function_stats.get(func, 0) + 1

            sheet_info = SheetInfo(
                name=sheet_name,
                used_range=ws.dimensions or "A1",
                row_count=row_count,
                column_count=col_count,
                formula_count=formula_count,
                cell_count=cell_count,
                has_tables=len(ws.tables) > 0 if hasattr(ws, 'tables') else False
            )
            result.sheets.append(sheet_info)

            # Extract tables
            if hasattr(ws, 'tables'):
                for table_name, table in ws.tables.items():
                    table_info = TableInfo(
                        name=table_name,
                        sheet=sheet_name,
                        range=table.ref,
                        row_count=0  # Would need to parse range
                    )
                    result.tables.append(table_info)

        # Extract named ranges
        if wb.defined_names:
            for name in wb.defined_names.definedName:
                named_range = NamedRangeInfo(
                    name=name.name,
                    scope='workbook' if name.localSheetId is None else wb.sheetnames[name.localSheetId],
                    refers_to=name.attr_text,
                    is_formula='(' in name.attr_text if name.attr_text else False
                )
                result.named_ranges.append(named_range)

        wb.close()

    def _analyze_xls(self, result: AnalysisResult):
        """Analyze .xls files using xlrd."""
        if xlrd is None:
            result.errors.append("xlrd not installed. Run: pip install xlrd")
            return

        try:
            wb = xlrd.open_workbook(self.file_path, formatting_info=False)
        except Exception as e:
            result.errors.append(f"Failed to open workbook: {str(e)}")
            return

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)

            sheet_info = SheetInfo(
                name=ws.name,
                used_range=f"A1:{get_column_letter(ws.ncols)}{ws.nrows}",
                row_count=ws.nrows,
                column_count=ws.ncols,
                formula_count=0,  # xlrd doesn't easily expose formulas
                cell_count=ws.nrows * ws.ncols
            )
            result.sheets.append(sheet_info)

        # Note: xlrd has limited formula support
        result.errors.append("Note: .xls format has limited formula extraction. Consider converting to .xlsx")

    def _analyze_csv(self, result: AnalysisResult):
        """Analyze CSV files."""
        import csv

        try:
            with open(self.file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                rows = list(reader)
        except Exception as e:
            result.errors.append(f"Failed to read CSV: {str(e)}")
            return

        row_count = len(rows)
        col_count = max(len(row) for row in rows) if rows else 0

        sheet_info = SheetInfo(
            name="Sheet1",
            used_range=f"A1:{get_column_letter(col_count) if col_count else 'A'}{row_count}",
            row_count=row_count,
            column_count=col_count,
            formula_count=0,
            cell_count=sum(len(row) for row in rows)
        )
        result.sheets.append(sheet_info)

        result.errors.append("Note: CSV files do not contain formulas")


def get_column_letter(col_idx: int) -> str:
    """Convert column index to letter (1=A, 2=B, etc.)."""
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result or "A"


# Convenience function
def analyze_spreadsheet(file_path: str) -> AnalysisResult:
    """Analyze a spreadsheet and return results."""
    analyzer = SpreadsheetAnalyzer()
    return analyzer.analyze(file_path)


# ============================================================================
# FORMULA EDITOR
# ============================================================================

@dataclass
class FormulaChange:
    """Represents a single formula change."""
    sheet: str
    address: str
    old_formula: str
    new_formula: str
    applied: bool = False
    error: Optional[str] = None


@dataclass
class EditResult:
    """Result of a formula edit operation."""
    success: bool
    changes_made: int
    changes: List[FormulaChange] = field(default_factory=list)
    backup_path: Optional[str] = None
    errors: List[str] = field(default_factory=list)


class SpreadsheetEditor:
    """Edit formulas across multiple sheets in Excel files."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self._load_workbook()

    def _load_workbook(self):
        """Load the workbook for editing."""
        if openpyxl is None:
            raise ImportError("openpyxl is required for editing. Run: pip install openpyxl")

        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        suffix = self.file_path.suffix.lower()
        if suffix not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Only .xlsx and .xlsm files can be edited. Got: {suffix}")

        self.workbook = openpyxl.load_workbook(
            self.file_path,
            data_only=False
        )

    def get_all_formulas(self) -> List[FormulaInfo]:
        """Get all formulas in the workbook."""
        formulas = []

        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                        formula_str = str(cell.value)
                        if formula_str.startswith('='):
                            formula_str = formula_str[1:]

                        formulas.append(FormulaInfo(
                            address=cell.coordinate,
                            sheet=sheet_name,
                            formula=formula_str,
                            functions=extract_functions(formula_str),
                            dependencies=extract_cell_references(formula_str),
                            is_array_formula=is_array_formula(formula_str)
                        ))

        return formulas

    def find_formulas(self, search_text: str, case_sensitive: bool = False,
                      use_regex: bool = False, sheets: Optional[List[str]] = None) -> List[FormulaInfo]:
        """Find formulas matching the search criteria."""
        all_formulas = self.get_all_formulas()
        matches = []

        for formula in all_formulas:
            # Filter by sheets if specified
            if sheets and formula.sheet not in sheets:
                continue

            formula_text = formula.formula

            if use_regex:
                try:
                    flags = 0 if case_sensitive else re.IGNORECASE
                    if re.search(search_text, formula_text, flags):
                        matches.append(formula)
                except re.error:
                    continue
            else:
                if case_sensitive:
                    if search_text in formula_text:
                        matches.append(formula)
                else:
                    if search_text.lower() in formula_text.lower():
                        matches.append(formula)

        return matches

    def preview_replace(self, find_text: str, replace_text: str,
                        case_sensitive: bool = False, use_regex: bool = False,
                        sheets: Optional[List[str]] = None) -> List[FormulaChange]:
        """Preview formula replacements without applying them."""
        matches = self.find_formulas(find_text, case_sensitive, use_regex, sheets)
        changes = []

        for formula in matches:
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                new_formula = re.sub(find_text, replace_text, formula.formula, flags=flags)
            else:
                if case_sensitive:
                    new_formula = formula.formula.replace(find_text, replace_text)
                else:
                    # Case-insensitive replace
                    pattern = re.compile(re.escape(find_text), re.IGNORECASE)
                    new_formula = pattern.sub(replace_text, formula.formula)

            if new_formula != formula.formula:
                changes.append(FormulaChange(
                    sheet=formula.sheet,
                    address=formula.address,
                    old_formula=formula.formula,
                    new_formula=new_formula
                ))

        return changes

    def apply_changes(self, changes: List[FormulaChange], create_backup: bool = True) -> EditResult:
        """Apply formula changes to the workbook."""
        result = EditResult(success=True, changes_made=0)

        # Create backup
        if create_backup:
            backup_path = self.file_path.with_suffix(f'.backup{self.file_path.suffix}')
            try:
                import shutil
                shutil.copy2(self.file_path, backup_path)
                result.backup_path = str(backup_path)
            except Exception as e:
                result.errors.append(f"Failed to create backup: {str(e)}")

        # Apply changes
        for change in changes:
            try:
                ws = self.workbook[change.sheet]
                cell = ws[change.address]
                cell.value = f"={change.new_formula}"
                change.applied = True
                result.changes_made += 1
            except Exception as e:
                change.error = str(e)
                result.errors.append(f"Failed to update {change.sheet}!{change.address}: {str(e)}")

        result.changes = changes
        return result

    def save(self, output_path: Optional[str] = None) -> str:
        """Save the workbook."""
        save_path = Path(output_path) if output_path else self.file_path

        try:
            self.workbook.save(save_path)
            return str(save_path)
        except Exception as e:
            raise IOError(f"Failed to save workbook: {str(e)}")

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()

    def replace_formulas(self, find_text: str, replace_text: str,
                         case_sensitive: bool = False, use_regex: bool = False,
                         sheets: Optional[List[str]] = None,
                         create_backup: bool = True,
                         output_path: Optional[str] = None) -> EditResult:
        """Find and replace formulas, then save."""
        # Preview changes
        changes = self.preview_replace(find_text, replace_text, case_sensitive, use_regex, sheets)

        if not changes:
            return EditResult(success=True, changes_made=0, errors=["No matching formulas found"])

        # Apply changes
        result = self.apply_changes(changes, create_backup)

        # Save
        if result.changes_made > 0:
            try:
                saved_path = self.save(output_path)
                result.errors.append(f"Saved to: {saved_path}")
            except IOError as e:
                result.success = False
                result.errors.append(str(e))

        return result

    def batch_replace(self, replacements: List[Dict[str, str]],
                      create_backup: bool = True,
                      output_path: Optional[str] = None) -> EditResult:
        """
        Apply multiple find/replace operations.

        Args:
            replacements: List of dicts with 'find', 'replace', and optional 'sheets', 'case_sensitive', 'use_regex'
        """
        all_changes = []

        for repl in replacements:
            find_text = repl.get('find', '')
            replace_text = repl.get('replace', '')
            case_sensitive = repl.get('case_sensitive', False)
            use_regex = repl.get('use_regex', False)
            sheets = repl.get('sheets', None)

            if find_text:
                changes = self.preview_replace(find_text, replace_text, case_sensitive, use_regex, sheets)
                all_changes.extend(changes)

        if not all_changes:
            return EditResult(success=True, changes_made=0, errors=["No matching formulas found"])

        # Apply all changes
        result = self.apply_changes(all_changes, create_backup)

        # Save
        if result.changes_made > 0:
            try:
                saved_path = self.save(output_path)
                result.errors.append(f"Saved to: {saved_path}")
            except IOError as e:
                result.success = False
                result.errors.append(str(e))

        return result

    def rename_sheet_references(self, old_sheet_name: str, new_sheet_name: str,
                                create_backup: bool = True) -> EditResult:
        """Rename sheet references in all formulas."""
        # Pattern to match sheet references like 'Sheet1'!A1 or Sheet1!A1
        find_patterns = [
            (f"'{old_sheet_name}'!", f"'{new_sheet_name}'!"),
            (f"{old_sheet_name}!", f"{new_sheet_name}!"),
        ]

        all_changes = []
        for find_text, replace_text in find_patterns:
            changes = self.preview_replace(find_text, replace_text, case_sensitive=True)
            all_changes.extend(changes)

        if not all_changes:
            return EditResult(success=True, changes_made=0, errors=["No sheet references found"])

        result = self.apply_changes(all_changes, create_backup)

        if result.changes_made > 0:
            try:
                self.save()
            except IOError as e:
                result.success = False
                result.errors.append(str(e))

        return result

    def update_cell_references(self, old_ref: str, new_ref: str,
                               create_backup: bool = True) -> EditResult:
        """Update cell references across all formulas."""
        return self.replace_formulas(
            find_text=old_ref,
            replace_text=new_ref,
            case_sensitive=False,
            use_regex=False,
            create_backup=create_backup
        )


# Convenience functions for editing
def find_and_replace_formulas(file_path: str, find_text: str, replace_text: str,
                               output_path: Optional[str] = None, **kwargs) -> EditResult:
    """Find and replace formulas in a spreadsheet."""
    editor = SpreadsheetEditor(file_path)
    try:
        return editor.replace_formulas(find_text, replace_text, output_path=output_path, **kwargs)
    finally:
        editor.close()


def preview_formula_changes(file_path: str, find_text: str, replace_text: str, **kwargs) -> List[FormulaChange]:
    """Preview formula changes without applying them."""
    editor = SpreadsheetEditor(file_path)
    try:
        return editor.preview_replace(find_text, replace_text, **kwargs)
    finally:
        editor.close()


# ============================================================================
# FORMULA TO CODE CONVERTER
# ============================================================================

@dataclass
class CodeConversion:
    """Result of converting a formula to code."""
    original_formula: str
    python_code: str
    javascript_code: str
    cell_address: Optional[str] = None
    sheet_name: Optional[str] = None
    variables_used: List[str] = field(default_factory=list)
    functions_used: List[str] = field(default_factory=list)
    notes: List[str] = field(default_factory=list)


class FormulaToCodeConverter:
    """Convert Excel formulas to Python and JavaScript code."""

    # Mapping of Excel functions to Python equivalents
    PYTHON_FUNCTIONS = {
        # Math
        'ABS': 'abs',
        'ROUND': 'round',
        'ROUNDUP': 'math.ceil',
        'ROUNDDOWN': 'math.floor',
        'INT': 'int',
        'MOD': 'lambda a, b: a % b',
        'POWER': 'pow',
        'SQRT': 'math.sqrt',
        'SUM': 'sum',
        'AVERAGE': 'lambda *args: sum(args) / len(args)',
        'MIN': 'min',
        'MAX': 'max',
        'CEILING': 'math.ceil',
        'FLOOR': 'math.floor',
        'RAND': 'random.random',
        'RANDBETWEEN': 'random.randint',

        # Logical
        'IF': 'lambda cond, true_val, false_val: true_val if cond else false_val',
        'AND': 'all',
        'OR': 'any',
        'NOT': 'not',
        'TRUE': 'True',
        'FALSE': 'False',
        'IFERROR': 'lambda val, fallback: fallback if isinstance(val, Exception) else val',
        'IFS': '# Use chained if/elif/else',
        'SWITCH': '# Use dict.get() or match/case',

        # Text
        'LEN': 'len',
        'LEFT': 'lambda s, n: s[:n]',
        'RIGHT': 'lambda s, n: s[-n:]',
        'MID': 'lambda s, start, length: s[start-1:start-1+length]',
        'UPPER': 'str.upper',
        'LOWER': 'str.lower',
        'PROPER': 'str.title',
        'TRIM': 'str.strip',
        'SUBSTITUTE': 'str.replace',
        'CONCATENATE': "lambda *args: ''.join(str(a) for a in args)",
        'TEXT': 'format',
        'VALUE': 'float',
        'FIND': 'str.find',
        'SEARCH': 'str.lower().find',
        'REPLACE': 'lambda s, start, num, new: s[:start-1] + new + s[start-1+num:]',

        # Lookup
        'VLOOKUP': '# Use dict lookup or pandas.merge',
        'HLOOKUP': '# Use dict lookup or pandas',
        'XLOOKUP': '# Use dict.get() or next() with filter',
        'INDEX': 'lambda arr, row, col=None: arr[row-1] if col is None else arr[row-1][col-1]',
        'MATCH': 'lambda val, arr, _: arr.index(val) + 1',
        'OFFSET': '# Calculate cell reference dynamically',
        'INDIRECT': '# Dynamic cell reference - use dict or getattr',
        'CHOOSE': 'lambda idx, *args: args[idx-1]',

        # Date/Time
        'TODAY': 'datetime.date.today',
        'NOW': 'datetime.datetime.now',
        'DATE': 'datetime.date',
        'YEAR': 'lambda d: d.year',
        'MONTH': 'lambda d: d.month',
        'DAY': 'lambda d: d.day',
        'HOUR': 'lambda d: d.hour',
        'MINUTE': 'lambda d: d.minute',
        'SECOND': 'lambda d: d.second',
        'DATEDIF': '# Calculate date difference manually',
        'EDATE': 'lambda d, months: d + relativedelta(months=months)',
        'EOMONTH': '# Use calendar.monthrange or dateutil',

        # Statistical
        'COUNT': 'len',
        'COUNTA': 'lambda arr: sum(1 for x in arr if x is not None)',
        'COUNTIF': 'lambda arr, cond: sum(1 for x in arr if cond(x))',
        'COUNTIFS': '# Use list comprehension with multiple conditions',
        'SUMIF': 'lambda arr, cond, sum_arr=None: sum(x for x in (sum_arr or arr) if cond(x))',
        'SUMIFS': '# Use list comprehension with multiple conditions',
        'AVERAGEIF': '# Use list comprehension with condition',

        # Array/Dynamic (Excel 365)
        'FILTER': 'lambda arr, cond: [x for x, c in zip(arr, cond) if c]',
        'SORT': 'sorted',
        'SORTBY': 'lambda arr, key_arr: [x for _, x in sorted(zip(key_arr, arr))]',
        'UNIQUE': 'lambda arr: list(dict.fromkeys(arr))',
        'SEQUENCE': 'lambda rows, cols=1, start=1, step=1: [[start + step*(r*cols+c) for c in range(cols)] for r in range(rows)]',
        'LET': '# Use intermediate variables',

        # Info
        'ISBLANK': 'lambda x: x is None or x == ""',
        'ISERROR': 'lambda x: isinstance(x, Exception)',
        'ISNUMBER': 'lambda x: isinstance(x, (int, float))',
        'ISTEXT': 'lambda x: isinstance(x, str)',
        'ISNA': 'lambda x: x is None',
    }

    # Mapping of Excel functions to JavaScript equivalents
    JS_FUNCTIONS = {
        # Math
        'ABS': 'Math.abs',
        'ROUND': 'Math.round',
        'ROUNDUP': 'Math.ceil',
        'ROUNDDOWN': 'Math.floor',
        'INT': 'Math.trunc',
        'MOD': '(a, b) => a % b',
        'POWER': 'Math.pow',
        'SQRT': 'Math.sqrt',
        'SUM': '(...args) => args.flat().reduce((a, b) => a + b, 0)',
        'AVERAGE': '(...args) => args.flat().reduce((a, b) => a + b, 0) / args.flat().length',
        'MIN': 'Math.min',
        'MAX': 'Math.max',
        'CEILING': 'Math.ceil',
        'FLOOR': 'Math.floor',
        'RAND': 'Math.random',
        'RANDBETWEEN': '(min, max) => Math.floor(Math.random() * (max - min + 1)) + min',

        # Logical
        'IF': '(cond, trueVal, falseVal) => cond ? trueVal : falseVal',
        'AND': '(...args) => args.every(Boolean)',
        'OR': '(...args) => args.some(Boolean)',
        'NOT': '!',
        'TRUE': 'true',
        'FALSE': 'false',
        'IFERROR': '(val, fallback) => { try { return val; } catch { return fallback; } }',

        # Text
        'LEN': 's => s.length',
        'LEFT': '(s, n) => s.slice(0, n)',
        'RIGHT': '(s, n) => s.slice(-n)',
        'MID': '(s, start, len) => s.substr(start - 1, len)',
        'UPPER': 's => s.toUpperCase()',
        'LOWER': 's => s.toLowerCase()',
        'TRIM': 's => s.trim()',
        'SUBSTITUTE': '(s, old, newStr) => s.replaceAll(old, newStr)',
        'CONCATENATE': '(...args) => args.join("")',
        'TEXT': '// Use Intl.NumberFormat or template literals',
        'VALUE': 'parseFloat',
        'FIND': '(find, text) => text.indexOf(find) + 1',
        'SEARCH': '(find, text) => text.toLowerCase().indexOf(find.toLowerCase()) + 1',

        # Lookup
        'VLOOKUP': '// Use array.find() or Map.get()',
        'XLOOKUP': '(lookup, arr, return_arr, fallback) => return_arr[arr.indexOf(lookup)] ?? fallback',
        'INDEX': '(arr, row, col) => col ? arr[row-1][col-1] : arr[row-1]',
        'MATCH': '(val, arr) => arr.indexOf(val) + 1',
        'CHOOSE': '(idx, ...args) => args[idx - 1]',

        # Date/Time
        'TODAY': '() => new Date().toISOString().split("T")[0]',
        'NOW': '() => new Date()',
        'YEAR': 'd => new Date(d).getFullYear()',
        'MONTH': 'd => new Date(d).getMonth() + 1',
        'DAY': 'd => new Date(d).getDate()',

        # Statistical
        'COUNT': 'arr => arr.filter(x => typeof x === "number").length',
        'COUNTA': 'arr => arr.filter(x => x != null && x !== "").length',
        'COUNTIF': '(arr, cond) => arr.filter(cond).length',
        'SUMIF': '(arr, cond, sumArr) => (sumArr || arr).filter((_, i) => cond(arr[i])).reduce((a, b) => a + b, 0)',

        # Array/Dynamic
        'FILTER': '(arr, cond) => arr.filter(cond)',
        'SORT': 'arr => [...arr].sort((a, b) => a - b)',
        'SORTBY': '(arr, keyArr) => arr.map((v, i) => [keyArr[i], v]).sort((a, b) => a[0] - b[0]).map(x => x[1])',
        'UNIQUE': 'arr => [...new Set(arr)]',

        # Info
        'ISBLANK': 'x => x == null || x === ""',
        'ISERROR': 'x => x instanceof Error',
        'ISNUMBER': 'x => typeof x === "number" && !isNaN(x)',
        'ISTEXT': 'x => typeof x === "string"',
    }

    def __init__(self):
        self.cell_pattern = re.compile(r"([A-Z]+)(\d+)")
        self.range_pattern = re.compile(r"([A-Z]+\d+):([A-Z]+\d+)")

    def convert_formula(self, formula: str, cell_address: str = None,
                        sheet_name: str = None) -> CodeConversion:
        """Convert an Excel formula to Python and JavaScript code."""
        # Remove leading = if present
        if formula.startswith('='):
            formula = formula[1:]

        functions_used = extract_functions(formula)
        cell_refs = extract_cell_references(formula)

        # Generate variable names from cell references
        variables = self._generate_variable_names(cell_refs)

        # Convert to Python
        python_code = self._convert_to_python(formula, variables, functions_used)

        # Convert to JavaScript
        js_code = self._convert_to_javascript(formula, variables, functions_used)

        notes = []

        # Add notes for complex functions
        for func in functions_used:
            if func.upper() in ['VLOOKUP', 'XLOOKUP', 'OFFSET', 'INDIRECT']:
                notes.append(f"{func}: Consider using a lookup table (dict/Map) for better performance")
            if func.upper() in ['SUMIFS', 'COUNTIFS', 'AVERAGEIFS']:
                notes.append(f"{func}: Use filtered aggregation with multiple conditions")
            if func.upper() in ['LET', 'LAMBDA']:
                notes.append(f"{func}: Use named intermediate variables for clarity")

        return CodeConversion(
            original_formula=formula,
            python_code=python_code,
            javascript_code=js_code,
            cell_address=cell_address,
            sheet_name=sheet_name,
            variables_used=list(variables.values()),
            functions_used=functions_used,
            notes=notes
        )

    def _generate_variable_names(self, cell_refs: List[str]) -> Dict[str, str]:
        """Generate meaningful variable names from cell references."""
        variables = {}
        for ref in cell_refs:
            # Clean up the reference
            clean_ref = ref.replace('$', '').replace("'", "")

            # Handle sheet references
            if '!' in clean_ref:
                sheet, cell = clean_ref.split('!')
                var_name = f"{self._sanitize_name(sheet)}_{cell.lower()}"
            else:
                var_name = f"cell_{clean_ref.lower()}"

            # Handle ranges
            if ':' in var_name:
                var_name = var_name.replace(':', '_to_')
                var_name = f"range_{var_name}"

            variables[ref] = var_name

        return variables

    def _sanitize_name(self, name: str) -> str:
        """Sanitize a name for use as a variable."""
        # Remove invalid characters
        name = re.sub(r'[^a-zA-Z0-9_]', '_', name)
        # Ensure it doesn't start with a number
        if name and name[0].isdigit():
            name = '_' + name
        return name.lower()

    def _convert_to_python(self, formula: str, variables: Dict[str, str],
                           functions: List[str]) -> str:
        """Convert formula to Python code."""
        code = formula

        # Replace cell references with variable names
        for ref, var in sorted(variables.items(), key=lambda x: -len(x[0])):
            code = code.replace(ref, var)

        # Replace operators
        code = code.replace('<>', '!=')
        code = code.replace('^', '**')
        code = code.replace('&', ' + ')  # String concatenation

        # Replace functions
        for func in functions:
            upper_func = func.upper()
            if upper_func in self.PYTHON_FUNCTIONS:
                py_func = self.PYTHON_FUNCTIONS[upper_func]
                # Handle lambda definitions vs simple replacements
                if py_func.startswith('lambda') or py_func.startswith('#'):
                    code = self._add_function_comment(code, func, py_func, 'Python')
                else:
                    pattern = re.compile(rf'\b{func}\s*\(', re.IGNORECASE)
                    code = pattern.sub(f'{py_func}(', code)

        # Format as Python
        lines = []
        lines.append("# Python equivalent")
        lines.append("import math")
        if any(f.upper() in ['TODAY', 'NOW', 'DATE', 'YEAR', 'MONTH', 'DAY'] for f in functions):
            lines.append("import datetime")
        if any(f.upper() in ['RAND', 'RANDBETWEEN'] for f in functions):
            lines.append("import random")
        lines.append("")

        # Add variable definitions as comments
        if variables:
            lines.append("# Variables (replace with actual values):")
            for ref, var in variables.items():
                lines.append(f"# {var} = <value from {ref}>")
            lines.append("")

        lines.append(f"result = {code}")

        return '\n'.join(lines)

    def _convert_to_javascript(self, formula: str, variables: Dict[str, str],
                                functions: List[str]) -> str:
        """Convert formula to JavaScript code."""
        code = formula

        # Replace cell references with variable names
        for ref, var in sorted(variables.items(), key=lambda x: -len(x[0])):
            code = code.replace(ref, var)

        # Replace operators
        code = code.replace('<>', '!==')
        code = code.replace('=', '===').replace('====', '===')  # Handle == comparisons
        code = code.replace('^', '**')
        code = code.replace('&', ' + ')  # String concatenation

        # Replace functions
        for func in functions:
            upper_func = func.upper()
            if upper_func in self.JS_FUNCTIONS:
                js_func = self.JS_FUNCTIONS[upper_func]
                if js_func.startswith('//') or '=>' in js_func:
                    code = self._add_function_comment(code, func, js_func, 'JavaScript')
                else:
                    pattern = re.compile(rf'\b{func}\s*\(', re.IGNORECASE)
                    code = pattern.sub(f'{js_func}(', code)

        # Format as JavaScript
        lines = []
        lines.append("// JavaScript equivalent")
        lines.append("")

        # Add variable definitions as comments
        if variables:
            lines.append("// Variables (replace with actual values):")
            for ref, var in variables.items():
                lines.append(f"// const {var} = /* value from {ref} */;")
            lines.append("")

        lines.append(f"const result = {code};")

        return '\n'.join(lines)

    def _add_function_comment(self, code: str, func: str, replacement: str, lang: str) -> str:
        """Add a comment explaining a complex function replacement."""
        # Keep the original function but add note
        return f"/* {func}: {replacement} */\n{code}"

    def convert_formulas_batch(self, formulas: List[FormulaInfo]) -> List[CodeConversion]:
        """Convert multiple formulas to code."""
        return [
            self.convert_formula(f.formula, f.address, f.sheet)
            for f in formulas
        ]

    def generate_python_module(self, formulas: List[FormulaInfo],
                                module_name: str = "spreadsheet_logic") -> str:
        """Generate a complete Python module from spreadsheet formulas."""
        lines = []
        lines.append(f'"""')
        lines.append(f'Auto-generated from spreadsheet formulas.')
        lines.append(f'Module: {module_name}')
        lines.append(f'"""')
        lines.append('')
        lines.append('import math')
        lines.append('import datetime')
        lines.append('from typing import Any, List, Optional')
        lines.append('')
        lines.append('')

        # Group formulas by sheet
        by_sheet: Dict[str, List[FormulaInfo]] = {}
        for f in formulas:
            if f.sheet not in by_sheet:
                by_sheet[f.sheet] = []
            by_sheet[f.sheet].append(f)

        # Generate class for each sheet
        for sheet_name, sheet_formulas in by_sheet.items():
            class_name = self._sanitize_name(sheet_name).title().replace('_', '')
            if not class_name:
                class_name = 'Sheet'

            lines.append(f'class {class_name}Calculator:')
            lines.append(f'    """Calculations from sheet: {sheet_name}"""')
            lines.append('')
            lines.append('    def __init__(self, data: dict):')
            lines.append('        """Initialize with cell data dictionary."""')
            lines.append('        self.data = data')
            lines.append('')

            for formula in sheet_formulas:
                method_name = f"calculate_{formula.address.lower()}"
                conversion = self.convert_formula(formula.formula, formula.address, formula.sheet)

                lines.append(f'    def {method_name}(self):')
                lines.append(f'        """')
                lines.append(f'        Cell {formula.address}: ={formula.formula}')
                if conversion.functions_used:
                    lines.append(f'        Functions: {", ".join(conversion.functions_used)}')
                lines.append(f'        """')

                # Add variable fetches
                for ref, var in self._generate_variable_names(extract_cell_references(formula.formula)).items():
                    lines.append(f'        {var} = self.data.get("{ref}", 0)')

                # Add simplified calculation
                lines.append(f'        # Original: ={formula.formula}')
                lines.append(f'        result = None  # TODO: Implement')
                lines.append(f'        return result')
                lines.append('')

            lines.append('')

        return '\n'.join(lines)


# Convenience function
def convert_formula_to_code(formula: str, cell_address: str = None) -> CodeConversion:
    """Convert a single formula to Python and JavaScript."""
    converter = FormulaToCodeConverter()
    return converter.convert_formula(formula, cell_address)


# ============================================================================
# DATA DICTIONARY GENERATOR
# ============================================================================

@dataclass
class DataDictionaryEntry:
    """Single entry in the data dictionary."""
    name: str
    entry_type: str  # 'named_range', 'table', 'column', 'sheet', 'formula_pattern'
    location: str
    description: str = ""
    data_type: str = ""
    formula: Optional[str] = None
    sample_values: List[str] = field(default_factory=list)
    dependencies: List[str] = field(default_factory=list)
    used_by: List[str] = field(default_factory=list)


@dataclass
class DataDictionary:
    """Complete data dictionary for a spreadsheet."""
    file_name: str
    generated_at: str
    entries: List[DataDictionaryEntry] = field(default_factory=list)
    summary: Dict[str, int] = field(default_factory=dict)


class DataDictionaryGenerator:
    """Generate data dictionary documentation from spreadsheets."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.workbook = None
        self._load_workbook()

    def _load_workbook(self):
        """Load the workbook for analysis."""
        if openpyxl is None:
            raise ImportError("openpyxl is required. Run: pip install openpyxl")

        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {self.file_path}")

        suffix = self.file_path.suffix.lower()
        if suffix not in ['.xlsx', '.xlsm', '.xlsb']:
            raise ValueError(f"Only Excel files supported. Got: {suffix}")

        self.workbook = openpyxl.load_workbook(
            self.file_path,
            data_only=False
        )

    def generate(self) -> DataDictionary:
        """Generate a complete data dictionary."""
        from datetime import datetime

        dictionary = DataDictionary(
            file_name=self.file_path.name,
            generated_at=datetime.now().isoformat()
        )

        # Extract named ranges
        self._extract_named_ranges(dictionary)

        # Extract tables
        self._extract_tables(dictionary)

        # Extract sheet info
        self._extract_sheets(dictionary)

        # Extract formula patterns
        self._extract_formula_patterns(dictionary)

        # Build dependency graph
        self._build_dependencies(dictionary)

        # Generate summary
        dictionary.summary = {
            'named_ranges': sum(1 for e in dictionary.entries if e.entry_type == 'named_range'),
            'tables': sum(1 for e in dictionary.entries if e.entry_type == 'table'),
            'columns': sum(1 for e in dictionary.entries if e.entry_type == 'column'),
            'sheets': sum(1 for e in dictionary.entries if e.entry_type == 'sheet'),
            'formula_patterns': sum(1 for e in dictionary.entries if e.entry_type == 'formula_pattern'),
            'total_entries': len(dictionary.entries)
        }

        return dictionary

    def _extract_named_ranges(self, dictionary: DataDictionary):
        """Extract all named ranges."""
        if not self.workbook.defined_names:
            return

        for name in self.workbook.defined_names.definedName:
            scope = 'workbook'
            if name.localSheetId is not None:
                scope = self.workbook.sheetnames[name.localSheetId]

            entry = DataDictionaryEntry(
                name=name.name,
                entry_type='named_range',
                location=name.attr_text or '',
                description=f"Named range in {scope}",
                formula=name.attr_text if name.attr_text and '(' in name.attr_text else None
            )

            # Try to get sample values
            try:
                if name.attr_text and '!' in name.attr_text:
                    # Parse the reference
                    ref = name.attr_text.replace('=', '')
                    if ':' in ref:
                        # Range reference
                        entry.description = f"Range: {ref}"
                    else:
                        # Single cell - try to get value
                        parts = ref.split('!')
                        if len(parts) == 2:
                            sheet_name = parts[0].strip("'")
                            cell_ref = parts[1]
                            if sheet_name in self.workbook.sheetnames:
                                ws = self.workbook[sheet_name]
                                cell = ws[cell_ref]
                                if cell.value is not None:
                                    entry.sample_values = [str(cell.value)[:50]]
            except Exception:
                pass

            dictionary.entries.append(entry)

    def _extract_tables(self, dictionary: DataDictionary):
        """Extract all tables and their columns."""
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            if not hasattr(ws, 'tables'):
                continue

            for table_name, table in ws.tables.items():
                # Table entry
                entry = DataDictionaryEntry(
                    name=table_name,
                    entry_type='table',
                    location=f"{sheet_name}!{table.ref}",
                    description=f"Excel table in {sheet_name}"
                )

                # Get headers and row count
                try:
                    from openpyxl.utils import range_boundaries
                    min_col, min_row, max_col, max_row = range_boundaries(table.ref)

                    headers = []
                    for col in range(min_col, max_col + 1):
                        cell = ws.cell(row=min_row, column=col)
                        if cell.value:
                            headers.append(str(cell.value))

                    entry.sample_values = headers
                    entry.description = f"Table with {max_row - min_row} rows, {len(headers)} columns"

                    # Add column entries
                    for i, header in enumerate(headers):
                        col_entry = DataDictionaryEntry(
                            name=f"{table_name}[{header}]",
                            entry_type='column',
                            location=f"{sheet_name}!{table.ref}",
                            description=f"Column in table {table_name}",
                            data_type=self._infer_column_type(ws, min_col + i, min_row + 1, max_row)
                        )

                        # Get sample values
                        samples = []
                        for row in range(min_row + 1, min(min_row + 4, max_row + 1)):
                            cell = ws.cell(row=row, column=min_col + i)
                            if cell.value is not None:
                                samples.append(str(cell.value)[:30])
                        col_entry.sample_values = samples

                        dictionary.entries.append(col_entry)

                except Exception:
                    pass

                dictionary.entries.append(entry)

    def _infer_column_type(self, ws, col: int, start_row: int, end_row: int) -> str:
        """Infer the data type of a column."""
        types_found = set()

        for row in range(start_row, min(start_row + 10, end_row + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue

            if isinstance(cell.value, bool):
                types_found.add('boolean')
            elif isinstance(cell.value, (int, float)):
                types_found.add('number')
            elif isinstance(cell.value, str):
                if cell.value.startswith('='):
                    types_found.add('formula')
                else:
                    types_found.add('text')
            elif hasattr(cell.value, 'strftime'):
                types_found.add('date')

        if len(types_found) == 0:
            return 'unknown'
        elif len(types_found) == 1:
            return types_found.pop()
        else:
            return 'mixed: ' + ', '.join(sorted(types_found))

    def _extract_sheets(self, dictionary: DataDictionary):
        """Extract sheet-level information."""
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            entry = DataDictionaryEntry(
                name=sheet_name,
                entry_type='sheet',
                location=sheet_name,
                description=f"Worksheet"
            )

            # Add dimensions
            if ws.dimensions:
                entry.description = f"Worksheet: {ws.dimensions}"

            # Count formulas
            formula_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                        formula_count += 1

            if formula_count:
                entry.description += f", {formula_count} formulas"

            dictionary.entries.append(entry)

    def _extract_formula_patterns(self, dictionary: DataDictionary):
        """Extract common formula patterns."""
        formula_patterns: Dict[str, List[str]] = {}

        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' or (isinstance(cell.value, str) and str(cell.value).startswith('=')):
                        formula = str(cell.value)
                        if formula.startswith('='):
                            formula = formula[1:]

                        # Extract the function pattern
                        functions = extract_functions(formula)
                        if functions:
                            pattern = '+'.join(sorted(set(f.upper() for f in functions)))
                            if pattern not in formula_patterns:
                                formula_patterns[pattern] = []
                            formula_patterns[pattern].append(f"{sheet_name}!{cell.coordinate}")

        # Add common patterns to dictionary
        for pattern, locations in sorted(formula_patterns.items(), key=lambda x: -len(x[1])):
            if len(locations) >= 2:  # Only include patterns used multiple times
                entry = DataDictionaryEntry(
                    name=f"Pattern: {pattern}",
                    entry_type='formula_pattern',
                    location=f"{len(locations)} cells",
                    description=f"Formula pattern using: {pattern.replace('+', ', ')}",
                    sample_values=locations[:5]  # First 5 locations
                )
                dictionary.entries.append(entry)

    def _build_dependencies(self, dictionary: DataDictionary):
        """Build dependency relationships between entries."""
        # Create name->entry lookup
        name_lookup = {e.name: e for e in dictionary.entries}

        for entry in dictionary.entries:
            if entry.entry_type == 'named_range' and entry.location:
                # Find what uses this named range
                for other in dictionary.entries:
                    if other.formula and entry.name in other.formula:
                        entry.used_by.append(other.name)
                        other.dependencies.append(entry.name)

    def export_markdown(self) -> str:
        """Export data dictionary as Markdown."""
        dd = self.generate()

        lines = []
        lines.append(f"# Data Dictionary: {dd.file_name}")
        lines.append("")
        lines.append(f"Generated: {dd.generated_at}")
        lines.append("")

        # Summary
        lines.append("## Summary")
        lines.append("")
        lines.append("| Category | Count |")
        lines.append("|----------|-------|")
        for key, value in dd.summary.items():
            lines.append(f"| {key.replace('_', ' ').title()} | {value} |")
        lines.append("")

        # Group by type
        by_type: Dict[str, List[DataDictionaryEntry]] = {}
        for entry in dd.entries:
            if entry.entry_type not in by_type:
                by_type[entry.entry_type] = []
            by_type[entry.entry_type].append(entry)

        # Named Ranges
        if 'named_range' in by_type:
            lines.append("## Named Ranges")
            lines.append("")
            lines.append("| Name | Location | Description |")
            lines.append("|------|----------|-------------|")
            for entry in by_type['named_range']:
                desc = entry.description
                if entry.sample_values:
                    desc += f" (e.g., {entry.sample_values[0]})"
                lines.append(f"| `{entry.name}` | {entry.location} | {desc} |")
            lines.append("")

        # Tables
        if 'table' in by_type:
            lines.append("## Tables")
            lines.append("")
            for entry in by_type['table']:
                lines.append(f"### {entry.name}")
                lines.append("")
                lines.append(f"- **Location**: {entry.location}")
                lines.append(f"- **Description**: {entry.description}")
                if entry.sample_values:
                    lines.append(f"- **Columns**: {', '.join(entry.sample_values)}")
                lines.append("")

        # Columns
        if 'column' in by_type:
            lines.append("## Table Columns")
            lines.append("")
            lines.append("| Column | Type | Samples |")
            lines.append("|--------|------|---------|")
            for entry in by_type['column']:
                samples = ', '.join(entry.sample_values[:3]) if entry.sample_values else '-'
                lines.append(f"| `{entry.name}` | {entry.data_type or '-'} | {samples} |")
            lines.append("")

        # Sheets
        if 'sheet' in by_type:
            lines.append("## Sheets")
            lines.append("")
            lines.append("| Sheet | Description |")
            lines.append("|-------|-------------|")
            for entry in by_type['sheet']:
                lines.append(f"| {entry.name} | {entry.description} |")
            lines.append("")

        # Formula Patterns
        if 'formula_pattern' in by_type:
            lines.append("## Common Formula Patterns")
            lines.append("")
            for entry in by_type['formula_pattern']:
                lines.append(f"### {entry.name}")
                lines.append(f"- Used in: {entry.location}")
                if entry.sample_values:
                    lines.append(f"- Examples: {', '.join(entry.sample_values)}")
                lines.append("")

        return '\n'.join(lines)

    def export_json(self) -> str:
        """Export data dictionary as JSON."""
        import json

        dd = self.generate()

        data = {
            'file_name': dd.file_name,
            'generated_at': dd.generated_at,
            'summary': dd.summary,
            'entries': [
                {
                    'name': e.name,
                    'type': e.entry_type,
                    'location': e.location,
                    'description': e.description,
                    'data_type': e.data_type,
                    'formula': e.formula,
                    'sample_values': e.sample_values,
                    'dependencies': e.dependencies,
                    'used_by': e.used_by
                }
                for e in dd.entries
            ]
        }

        return json.dumps(data, indent=2)

    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()


# Convenience functions
def generate_data_dictionary(file_path: str) -> DataDictionary:
    """Generate a data dictionary for a spreadsheet."""
    generator = DataDictionaryGenerator(file_path)
    try:
        return generator.generate()
    finally:
        generator.close()


def export_data_dictionary_markdown(file_path: str) -> str:
    """Export a data dictionary as Markdown."""
    generator = DataDictionaryGenerator(file_path)
    try:
        return generator.export_markdown()
    finally:
        generator.close()
