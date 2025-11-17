#!/usr/bin/env python3
"""
HOLT HOMES BAT PRICING UPDATER
Specifically configured for: HOLT BAT OCTOBER 2025 9-29-25.xlsm
Updates pricing from updatetool_MarginUpdates to costsheet_UnconvertedPricing
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime


class HoltHomesUpdater:
    def __init__(self, file_path):
        """Initialize with Holt Homes BAT file"""
        self.file_path = file_path
        self.wb = None
        self.updates_applied = 0
        self.items_updated = 0
        
        # Column mappings for Holt BAT
        self.pricing_cols = {
            'zone': 1,          # A: Pricing Zone
            'category': 2,      # B: Product Category
            'minor_category': 3,# C: Minor Category
            'item_id': 4,       # D: Item ID
            'base_cost': 18     # R: Base Cost
        }
        
        # Price level sell price columns
        self.price_level_cols = {
            'PL01': {'margin': 19, 'sell': 20, 'actual': 21},
            'PL02': {'margin': 22, 'sell': 23, 'actual': 24},
            'PL03': {'margin': 25, 'sell': 26, 'actual': 27},
            'PL04': {'margin': 28, 'sell': 29, 'actual': 30},
            'PL05': {'margin': 31, 'sell': 32, 'actual': 33},
            'PL06': {'margin': 34, 'sell': 35, 'actual': 36},
            'PL07': {'margin': 37, 'sell': 38, 'actual': 39},
            'PL08': {'margin': 40, 'sell': 41, 'actual': 42},
            'PL09': {'margin': 43, 'sell': 44, 'actual': 45},
            'PL10': {'margin': 46, 'sell': 47, 'actual': 48},
            'PL11': {'margin': 49, 'sell': 50, 'actual': 51},
            'PL12': {'margin': 52, 'sell': 53, 'actual': 54}
        }
    
    def run_update(self):
        """Main update process"""
        print("\n" + "="*70)
        print("HOLT HOMES BAT PRICING UPDATER")
        print("="*70)
        print(f"File: {os.path.basename(self.file_path)}")
        print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("-"*70)
        
        # Load file
        if not self.load_file():
            return False
        
        # Get updates
        updates = self.get_margin_updates()
        if not updates:
            print("⚠️ No margin updates found")
            return False
        
        # Apply updates
        if not self.apply_updates(updates):
            return False
        
        # Save file
        return self.save_file()
    
    def load_file(self):
        """Load the Excel file - handles both .xlsx and .xlsm"""
        if not os.path.exists(self.file_path):
            print(f"❌ File not found: {self.file_path}")
            return False
        
        print(f"📂 Loading BAT file...")
        try:
            # Important: keep_vba=True for .xlsm files to preserve macros
            if self.file_path.endswith('.xlsm'):
                self.wb = openpyxl.load_workbook(self.file_path, keep_vba=True)
            else:
                self.wb = openpyxl.load_workbook(self.file_path)
            print(f"✔ File loaded successfully")
            return True
        except Exception as e:
            print(f"❌ Error loading file: {e}")
            return False
    
    def get_margin_updates(self):
        """Read margin updates from updatetool_MarginUpdates sheet"""
        # Check for the sheet with various possible names
        sheet_name = None
        possible_names = ['updatetool_MarginUpdates', 'MARGINS TO CHANGE', 
                         'Margin Updates', 'MarginUpdates', 'MARGIN CHANGES TO APPLY']
        
        for name in possible_names:
            if name in self.wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"❌ Margin updates sheet not found")
            print(f"   Available sheets: {', '.join(self.wb.sheetnames[:5])}")
            return []
        
        sheet = self.wb[sheet_name]
        print(f"📋 Reading from: {sheet_name}")
        
        updates = []
        
        # Find header row
        header_row = None
        for row in range(1, min(10, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and 'Pricing Zone' in str(cell_value):
                header_row = row
                break
        
        if not header_row:
            # Try default row 4 (common in your files)
            header_row = 4
        
        print(f"   Starting from row {header_row + 1}")
        
        # Read updates
        for row in range(header_row + 1, sheet.max_row + 1):
            zone = sheet.cell(row=row, column=1).value
            category = sheet.cell(row=row, column=2).value
            minor_cat = sheet.cell(row=row, column=3).value
            item_id = sheet.cell(row=row, column=4).value
            margin = sheet.cell(row=row, column=6).value  # New Margin %
            price_levels = sheet.cell(row=row, column=8).value  # Price Levels
            
            # Skip empty rows
            if not zone or not category or margin is None:
                continue
            
            # Convert margin
            if isinstance(margin, str):
                if '%' in margin:
                    margin = float(margin.strip('%')) / 100
                else:
                    try:
                        margin = float(margin)
                        if margin > 1:
                            margin = margin / 100
                    except:
                        continue
            elif margin > 1:
                margin = margin / 100
            
            # Parse price levels
            pl_list = self.parse_price_levels(str(price_levels) if price_levels else 'ALL')
            
            updates.append({
                'zone': str(zone).strip(),
                'category': str(category).strip(),
                'minor_category': str(minor_cat).strip() if minor_cat else '',
                'item_id': str(item_id).strip().upper() if item_id else 'ALL',
                'margin': margin,
                'price_levels': pl_list,
                'update_row': row
            })
        
        print(f"✔ Found {len(updates)} margin updates")
        
        # Show summary
        if updates:
            print("\n📊 Update Summary:")
            for i, update in enumerate(updates[:5], 1):  # Show first 5
                pl_text = ','.join(update['price_levels'][:3]) 
                if len(update['price_levels']) > 3:
                    pl_text += '...'
                print(f"   {i}. {update['category'][:25]} → {update['margin']:.1%} [{pl_text}]")
            if len(updates) > 5:
                print(f"   ... and {len(updates) - 5} more")
        
        return updates
    
    def parse_price_levels(self, pl_str):
        """Parse price level string"""
        pl_str = pl_str.upper().strip()
        
        if pl_str in ['ALL', '', 'NONE']:
            return list(self.price_level_cols.keys())
        
        result = []
        
        # Handle comma-separated
        if ',' in pl_str:
            for pl in pl_str.split(','):
                pl = pl.strip()
                if pl.startswith('PL') and len(pl) == 4:
                    if pl in self.price_level_cols:
                        result.append(pl)
                elif pl.isdigit():
                    pl_code = f'PL{int(pl):02d}'
                    if pl_code in self.price_level_cols:
                        result.append(pl_code)
        
        # Handle range
        elif '-' in pl_str:
            parts = pl_str.split('-')
            if len(parts) == 2:
                start = parts[0].replace('PL', '').strip()
                end = parts[1].replace('PL', '').strip()
                if start.isdigit() and end.isdigit():
                    for i in range(int(start), min(int(end) + 1, 13)):
                        pl_code = f'PL{i:02d}'
                        if pl_code in self.price_level_cols:
                            result.append(pl_code)
        
        # Single level
        else:
            if pl_str.startswith('PL'):
                if pl_str in self.price_level_cols:
                    result.append(pl_str)
            elif pl_str.isdigit():
                pl_code = f'PL{int(pl_str):02d}'
                if pl_code in self.price_level_cols:
                    result.append(pl_code)
        
        return result if result else list(self.price_level_cols.keys())
    
    def apply_updates(self, updates):
        """Apply updates to costsheet_UnconvertedPricing"""
        # Find pricing sheet
        sheet_name = None
        possible_names = ['costsheet_UnconvertedPricing', 'Unconverted', 
                         'UnconvertedPricing', 'Pricing']
        
        for name in possible_names:
            if name in self.wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"❌ Pricing sheet not found")
            return False
        
        sheet = self.wb[sheet_name]
        print(f"\n⚙️ Updating: {sheet_name}")
        print(f"   Processing {sheet.max_row - 1:,} pricing rows...")
        
        # Light green highlight for updated cells
        highlight = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        
        # Process each update
        for idx, update in enumerate(updates, 1):
            matches = 0
            
            # Show progress for large files
            if idx % 10 == 0:
                print(f"   Processing update {idx}/{len(updates)}...")
            
            # Check each row
            for row in range(2, sheet.max_row + 1):
                if self.row_matches(sheet, row, update):
                    # Get base cost
                    base_cost = sheet.cell(row=row, column=self.pricing_cols['base_cost']).value
                    
                    if base_cost and isinstance(base_cost, (int, float)) and base_cost > 0:
                        # Update price levels
                        for pl in update['price_levels']:
                            if pl not in self.price_level_cols:
                                continue
                            
                            cols = self.price_level_cols[pl]
                            
                            # Update margin
                            margin_cell = sheet.cell(row=row, column=cols['margin'])
                            margin_cell.value = update['margin']
                            margin_cell.number_format = '0.0%'
                            margin_cell.fill = highlight
                            
                            # Update sell price formula
                            sell_cell = sheet.cell(row=row, column=cols['sell'])
                            base_col = get_column_letter(self.pricing_cols['base_cost'])
                            margin_col = get_column_letter(cols['margin'])
                            sell_cell.value = f'={base_col}{row}/(1-{margin_col}{row})'
                            sell_cell.number_format = '$#,##0.00'
                            sell_cell.fill = highlight
                            
                            # Update actual margin formula
                            actual_cell = sheet.cell(row=row, column=cols['actual'])
                            sell_col = get_column_letter(cols['sell'])
                            actual_cell.value = f'=({sell_col}{row}-{base_col}{row})/{sell_col}{row}'
                            actual_cell.number_format = '0.0%'
                            
                            self.updates_applied += 3
                        
                        matches += 1
            
            self.items_updated += matches
            
            # Update status in margin sheet
            if 'update_row' in update:
                for name in ['updatetool_MarginUpdates', 'MARGINS TO CHANGE']:
                    if name in self.wb.sheetnames:
                        status_sheet = self.wb[name]
                        status_cell = status_sheet.cell(row=update['update_row'], column=9)
                        pl_text = ','.join(update['price_levels']) if len(update['price_levels']) < 12 else 'ALL'
                        status_cell.value = f"✔ {matches} items [{pl_text}]"
                        break
        
        print(f"\n✅ Update Results:")
        print(f"   Items updated: {self.items_updated:,}")
        print(f"   Cells modified: {self.updates_applied:,}")
        
        return True
    
    def row_matches(self, sheet, row, update):
        """Check if row matches update criteria"""
        row_zone = str(sheet.cell(row=row, column=self.pricing_cols['zone']).value or '').strip()
        row_category = str(sheet.cell(row=row, column=self.pricing_cols['category']).value or '').strip()
        row_minor = str(sheet.cell(row=row, column=self.pricing_cols['minor_category']).value or '').strip()
        row_item = str(sheet.cell(row=row, column=self.pricing_cols['item_id']).value or '').strip().upper()
        
        # Zone and category must match
        if row_zone != update['zone'] or row_category != update['category']:
            return False
        
        # Check minor category if specified
        if update['minor_category'] and row_minor != update['minor_category']:
            return False
        
        # Check item ID
        if update['item_id'] != 'ALL' and row_item != update['item_id']:
            return False
        
        return True
    
    def save_file(self):
        """Save the updated file"""
        # Create backup name with timestamp
        base_name = os.path.splitext(self.file_path)[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = f"{base_name}_UPDATED_{timestamp}.xlsm"
        
        print(f"\n💾 Saving updated file...")
        print(f"   Output: {os.path.basename(output_file)}")
        
        try:
            self.wb.save(output_file)
            print(f"✔ File saved successfully")
            
            print("\n" + "="*70)
            print("UPDATE COMPLETE!")
            print("="*70)
            print(f"Original: {os.path.basename(self.file_path)}")
            print(f"Updated:  {os.path.basename(output_file)}")
            print(f"Items:    {self.items_updated:,} items updated")
            print(f"Cells:    {self.updates_applied:,} cells modified")
            
            return True
        except Exception as e:
            print(f"❌ Error saving file: {e}")
            return False


def main():
    # Default Holt Homes file path - UPDATED TO LOCAL PATH
    default_path = r"C:\Users\corey.boser\Documents\Holt Monthly Pricing\HOLT BAT OCTOBER 2025 9-29-25.xlsm"
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = default_path
        
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        print("\nPlease provide the correct file path:")
        print("python holt_updater.py 'path\\to\\your\\file.xlsm'")
        sys.exit(1)
    
    # Run updater
    updater = HoltHomesUpdater(file_path)
    success = updater.run_update()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
