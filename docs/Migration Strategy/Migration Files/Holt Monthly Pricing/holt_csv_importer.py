#!/usr/bin/env python3
"""
HOLT HOMES CSV IMPORTER
Imports Price_Update.csv into BAT file and applies updates
"""

import csv
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import sys
from holt_updater import HoltHomesUpdater


class HoltCSVImporter:
    def __init__(self, csv_file, bat_file):
        """Initialize with CSV and BAT files"""
        self.csv_file = csv_file
        self.bat_file = bat_file
        self.wb = None
        
    def import_csv_to_bat(self):
        """Import CSV data into BAT file"""
        print("\n" + "="*70)
        print("HOLT HOMES CSV IMPORTER")
        print("="*70)
        print(f"CSV: {os.path.basename(self.csv_file)}")
        print(f"BAT: {os.path.basename(self.bat_file)}")
        print("-"*70)
        
        # Load BAT file
        print("📂 Loading BAT file...")
        try:
            if self.bat_file.endswith('.xlsm'):
                self.wb = openpyxl.load_workbook(self.bat_file, keep_vba=True)
            else:
                self.wb = openpyxl.load_workbook(self.bat_file)
            print("✔ BAT file loaded")
        except Exception as e:
            print(f"❌ Error loading BAT: {e}")
            return None
        
        # Read CSV
        print("📄 Reading CSV file...")
        updates = []
        try:
            with open(self.csv_file, 'r', newline='', encoding='utf-8-sig') as file:
                reader = csv.DictReader(file)
                for row in reader:
                    # Parse the row based on expected columns
                    zone = row.get('Pricing Zone', row.get('Zone', ''))
                    category = row.get('Product Category', row.get('Category', ''))
                    minor_cat = row.get('Minor Category', row.get('Minor', ''))
                    item_id = row.get('Item ID', row.get('Item', 'ALL'))
                    margin = row.get('New Margin', row.get('Margin', ''))
                    price_levels = row.get('Price Levels', row.get('Levels', 'ALL'))
                    
                    # Convert margin to decimal
                    if isinstance(margin, str):
                        margin = margin.strip()
                        if '%' in margin:
                            margin = float(margin.strip('%')) / 100
                        else:
                            margin = float(margin)
                            if margin > 1:
                                margin = margin / 100
                    
                    updates.append({
                        'zone': zone.strip(),
                        'category': category.strip(),
                        'minor_category': minor_cat.strip() if minor_cat else '',
                        'item_id': item_id.strip().upper() if item_id and item_id != 'ALL' else 'ALL',
                        'margin': margin,
                        'price_levels': price_levels.strip()
                    })
            
            print(f"✔ Found {len(updates)} updates in CSV")
            
        except Exception as e:
            print(f"❌ Error reading CSV: {e}")
            return None
        
        # Create or update margin sheet
        sheet_name = 'updatetool_MarginUpdates'
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            print(f"📋 Updating existing: {sheet_name}")
        else:
            sheet = self.wb.create_sheet(sheet_name, 0)
            print(f"📋 Creating new sheet: {sheet_name}")
        
        # Clear existing content
        sheet.delete_rows(1, sheet.max_row)
        
        # Add headers
        headers = ['Pricing Zone', 'Product Category', 'Minor Category', 
                  'Item ID', '', 'New Margin %', '', 'Price Levels', 'Status']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col)
            cell.value = header
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Add data
        for idx, update in enumerate(updates, 5):
            sheet.cell(row=idx, column=1).value = update['zone']
            sheet.cell(row=idx, column=2).value = update['category']
            sheet.cell(row=idx, column=3).value = update['minor_category']
            sheet.cell(row=idx, column=4).value = update['item_id']
            sheet.cell(row=idx, column=6).value = update['margin']
            sheet.cell(row=idx, column=6).number_format = '0.0%'
            sheet.cell(row=idx, column=8).value = update['price_levels']
            sheet.cell(row=idx, column=9).value = 'Pending'
        
        # Save with CSV marker
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        base_name = os.path.splitext(self.bat_file)[0]
        output_file = f"{base_name}_WITH_CSV_{timestamp}.xlsm"
        
        print(f"\n💾 Saving BAT with imported CSV data...")
        try:
            self.wb.save(output_file)
            print(f"✔ Saved: {os.path.basename(output_file)}")
            return output_file
        except Exception as e:
            print(f"❌ Error saving: {e}")
            return None


def main():
    if len(sys.argv) < 3:
        # Default paths
        csv_file = "Price_Update.csv"
        bat_file = r"C:\Users\corey.boser\Documents\Holt Monthly Pricing\HOLT BAT OCTOBER 2025 9-29-25.xlsm"
        
        # Check if CSV exists in current directory
        if not os.path.exists(csv_file):
            print(f"❌ CSV file not found: {csv_file}")
            print("\nUsage: python holt_csv_importer.py Price_Update.csv 'path\\to\\BAT.xlsm'")
            sys.exit(1)
    else:
        csv_file = sys.argv[1]
        bat_file = sys.argv[2]
    
    # Check files exist
    if not os.path.exists(csv_file):
        print(f"❌ CSV file not found: {csv_file}")
        sys.exit(1)
    
    if not os.path.exists(bat_file):
        print(f"❌ BAT file not found: {bat_file}")
        sys.exit(1)
    
    # Step 1: Import CSV
    importer = HoltCSVImporter(csv_file, bat_file)
    imported_file = importer.import_csv_to_bat()
    
    if not imported_file:
        print("❌ CSV import failed")
        sys.exit(1)
    
    print("\n" + "="*70)
    print("STEP 2: APPLYING UPDATES")
    print("="*70)
    
    # Step 2: Run updater on the imported file
    updater = HoltHomesUpdater(imported_file)
    success = updater.run_update()
    
    if success:
        print("\n" + "="*70)
        print("CSV IMPORT AND UPDATE COMPLETE!")
        print("="*70)
        print("Created two files:")
        print(f"1. {os.path.basename(imported_file)} (with CSV data)")
        print(f"2. {os.path.basename(imported_file.replace('WITH_CSV', 'UPDATED'))} (with applied pricing)")
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
