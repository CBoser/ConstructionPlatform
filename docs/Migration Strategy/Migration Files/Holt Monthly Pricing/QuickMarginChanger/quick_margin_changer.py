#!/usr/bin/env python3
"""
QUICK MARGIN CHANGER - Pricing Generator Tool
For use with QUICK_MARGIN_CHANGER.xlsm
Generates pricing data that can be copied to BAT files
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
from datetime import datetime


class QuickMarginChanger:
    def __init__(self, file_path):
        """Initialize with QUICK_MARGIN_CHANGER.xlsm file"""
        self.file_path = file_path
        self.wb = None
        self.output_wb = None
        self.updates_applied = 0
        self.items_updated = 0
        
        # Column mappings for margin input
        self.margin_cols = {
            'zone': 1,          # A: Pricing Zone
            'category': 2,      # B: Product Category
            'minor_category': 3,# C: Minor Category
            'item_id': 4,       # D: Item ID
            'description': 5,   # E: Description (optional)
            'margin': 6,        # F: New Margin %
            'notes': 7,         # G: Notes (optional)
            'price_levels': 8,  # H: Price Levels
            'status': 9         # I: Status
        }
        
        # Price level sell price columns (for output)
        self.price_level_cols = {
            'PL01': {'margin': 19, 'sell': 20, 'actual': 21},
            'PL02': {'margin': 22, 'sell': 23, 'actual': 24},
            'PL03': {'margin': 25, 'sell': 26, 'actual': 27},
            'PL04': {'margin': 28, 'sell': 29, 'actual': 30},
            'PL05': {'margin': 31, 'sell': 32, 'actual': 33},
            'PL06': {'margin': 34, 'sell': 35, 'actual': 36},
            'PL07': {'margin': 37, 'sell': 38, 'actual': 39},
            'PL08': {'margin': 40, 'sell': 41, 'actual': 42},
            'PL09': {'margin': 43, 'sell': 44, 'actual': 45},
            'PL10': {'margin': 46, 'sell': 47, 'actual': 48},
            'PL11': {'margin': 49, 'sell': 50, 'actual': 51},
            'PL12': {'margin': 52, 'sell': 53, 'actual': 54}
        }
    
    def run_update(self):
        """Main process to generate pricing"""
        print("\n" + "="*70)
        print("QUICK MARGIN CHANGER - Pricing Generator")
        print("="*70)
        print(f"File: {os.path.basename(self.file_path)}")
        print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("-"*70)
        
        # Load file
        if not self.load_file():
            return False
        
        # Get margin updates
        updates = self.get_margin_updates()
        if not updates:
            print("⚠️ No margin updates found")
            return False
        
        # Create pricing output
        if not self.generate_pricing(updates):
            return False
        
        # Save file
        return self.save_output()
    
    def load_file(self):
        """Load the Excel file"""
        if not os.path.exists(self.file_path):
            print(f"❌ File not found: {self.file_path}")
            return False
        
        print(f"📂 Loading QUICK_MARGIN_CHANGER file...")
        try:
            # Important: keep_vba=True for .xlsm files to preserve macros
            if self.file_path.endswith('.xlsm'):
                self.wb = openpyxl.load_workbook(self.file_path, keep_vba=True)
            else:
                self.wb = openpyxl.load_workbook(self.file_path)
            print(f"✔ File loaded successfully")
            return True
        except Exception as e:
            print(f"❌ Error loading file: {e}")
            return False
    
    def get_margin_updates(self):
        """Read margin updates from the input sheet"""
        # Check for the sheet with various possible names
        sheet_name = None
        possible_names = ['Margin Input', 'INPUT', 'MARGINS TO CHANGE', 
                         'Margin Updates', 'MARGIN INPUT']
        
        for name in possible_names:
            if name in self.wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            print(f"❌ Margin input sheet not found")
            print(f"   Available sheets: {', '.join(self.wb.sheetnames[:5])}")
            return []
        
        sheet = self.wb[sheet_name]
        print(f"📋 Reading from: {sheet_name}")
        
        updates = []
        
        # Find header row
        header_row = None
        for row in range(1, min(10, sheet.max_row + 1)):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value and 'Pricing Zone' in str(cell_value):
                header_row = row
                break
        
        if not header_row:
            # Try default row 4 (common in the BAT files)
            header_row = 4
        
        print(f"   Starting from row {header_row + 1}")
        
        # Read updates
        for row in range(header_row + 1, sheet.max_row + 1):
            zone = sheet.cell(row=row, column=self.margin_cols['zone']).value
            category = sheet.cell(row=row, column=self.margin_cols['category']).value
            minor_cat = sheet.cell(row=row, column=self.margin_cols['minor_category']).value
            item_id = sheet.cell(row=row, column=self.margin_cols['item_id']).value
            description = sheet.cell(row=row, column=self.margin_cols['description']).value
            margin = sheet.cell(row=row, column=self.margin_cols['margin']).value  # New Margin %
            notes = sheet.cell(row=row, column=self.margin_cols['notes']).value
            price_levels = sheet.cell(row=row, column=self.margin_cols['price_levels']).value  # Price Levels
            
            # Skip empty rows
            if not zone or not category or margin is None:
                continue
            
            # Convert margin
            if isinstance(margin, str):
                if '%' in margin:
                    margin = float(margin.strip('%')) / 100
                else:
                    try:
                        margin = float(margin)
                        if margin > 1:
                            margin = margin / 100
                    except:
                        continue
            elif margin > 1:
                margin = margin / 100
            
            # Parse price levels
            pl_list = self.parse_price_levels(str(price_levels) if price_levels else 'ALL')
            
            updates.append({
                'zone': str(zone).strip(),
                'category': str(category).strip(),
                'minor_category': str(minor_cat).strip() if minor_cat else '',
                'item_id': str(item_id).strip().upper() if item_id else 'ALL',
                'description': str(description).strip() if description else '',
                'margin': margin,
                'notes': str(notes).strip() if notes else '',
                'price_levels': pl_list,
                'update_row': row
            })
        
        print(f"✔ Found {len(updates)} margin updates")
        
        # Show summary
        if updates:
            print("\n📊 Update Summary:")
            for i, update in enumerate(updates[:5], 1):  # Show first 5
                pl_text = ','.join(update['price_levels'][:3]) 
                if len(update['price_levels']) > 3:
                    pl_text += '...'
                print(f"   {i}. {update['category'][:25]} → {update['margin']:.1%} [{pl_text}]")
            if len(updates) > 5:
                print(f"   ... and {len(updates) - 5} more")
        
        return updates
    
    def parse_price_levels(self, pl_str):
        """Parse price level string"""
        pl_str = pl_str.upper().strip()
        
        if pl_str in ['ALL', '', 'NONE']:
            return list(self.price_level_cols.keys())
        
        result = []
        
        # Handle comma-separated
        if ',' in pl_str:
            for pl in pl_str.split(','):
                pl = pl.strip()
                if pl.startswith('PL') and len(pl) == 4:
                    if pl in self.price_level_cols:
                        result.append(pl)
                elif pl.isdigit():
                    pl_code = f'PL{int(pl):02d}'
                    if pl_code in self.price_level_cols:
                        result.append(pl_code)
        
        # Handle range
        elif '-' in pl_str:
            parts = pl_str.split('-')
            if len(parts) == 2:
                start = parts[0].replace('PL', '').strip()
                end = parts[1].replace('PL', '').strip()
                if start.isdigit() and end.isdigit():
                    for i in range(int(start), min(int(end) + 1, 13)):
                        pl_code = f'PL{i:02d}'
                        if pl_code in self.price_level_cols:
                            result.append(pl_code)
        
        # Single level
        else:
            if pl_str.startswith('PL'):
                if pl_str in self.price_level_cols:
                    result.append(pl_str)
            elif pl_str.isdigit():
                pl_code = f'PL{int(pl_str):02d}'
                if pl_code in self.price_level_cols:
                    result.append(pl_code)
        
        return result if result else list(self.price_level_cols.keys())
    
    def generate_pricing(self, updates):
        """Generate pricing in new workbook for copy-paste"""
        print(f"\n⚙️ Generating pricing data...")
        
        # Create new workbook for output
        self.output_wb = openpyxl.Workbook()
        
        # Create output sheets for different BATs
        self.create_holt_bat_sheet(updates)
        self.create_richmond_bat_sheet(updates)
        
        # Create a lookup sheet
        self.create_lookup_sheet(updates)
        
        print(f"\n✅ Generation Results:")
        print(f"   Items processed: {len(updates):,}")
        print(f"   Output sheets created: {len(self.output_wb.sheetnames):,}")
        
        return True
        
    def create_holt_bat_sheet(self, updates):
        """Create Holt BAT format sheet"""
        # Use the first sheet and rename it
        sheet = self.output_wb.active
        sheet.title = "Holt BAT Format"
        
        # Setup headers
        headers = [
            "Pricing Zone", "Product Category", "Minor Category", "Item ID", 
            "Description", "New Margin", "Price Levels", "Notes"
        ]
        
        # Setup styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        # Adjust column widths
        col_widths = [15, 20, 20, 15, 30, 12, 15, 30]
        for i, width in enumerate(col_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
            
        # Fill data
        for row_idx, update in enumerate(updates, 2):
            # Format margin as percentage
            if isinstance(update['margin'], (int, float)):
                margin_display = f"{update['margin']:.1%}"
            else:
                margin_display = str(update['margin'])
                
            # Format price levels
            price_levels = ",".join(update['price_levels']) if len(update['price_levels']) < 12 else "ALL"
            
            # Write row
            sheet.cell(row=row_idx, column=1).value = update['zone']
            sheet.cell(row=row_idx, column=2).value = update['category']
            sheet.cell(row=row_idx, column=3).value = update['minor_category']
            sheet.cell(row=row_idx, column=4).value = update['item_id']
            sheet.cell(row=row_idx, column=5).value = update['description']
            sheet.cell(row=row_idx, column=6).value = margin_display
            sheet.cell(row=row_idx, column=7).value = price_levels
            sheet.cell(row=row_idx, column=8).value = update['notes']
            
            # Add light fill to margin cell
            sheet.cell(row=row_idx, column=6).fill = PatternFill(start_color="E0F0FF", end_color="E0F0FF", fill_type="solid")
            
            # Add borders
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col).border = border
        
        # Add instructions at the top
        instruction_row = sheet.insert_rows(1, 3)
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        
        cell = sheet.cell(row=1, column=1)
        cell.value = "HOLT BAT FORMAT - COPY AND PASTE TO UPDATETOOL_MARGINUPDATES SHEET"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        cell = sheet.cell(row=2, column=1)
        cell.value = "Instructions: Select all rows below (including headers) and paste to your Holt BAT"
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center')
        
        cell = sheet.cell(row=3, column=1)
        cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {len(updates)} items"
        cell.alignment = Alignment(horizontal='center')
        
    def create_richmond_bat_sheet(self, updates):
        """Create Richmond BAT format sheet"""
        # Create new sheet
        sheet = self.output_wb.create_sheet(title="Richmond BAT Format")
        
        # Setup headers
        headers = [
            "Pricing Zone", "Product Category", "Minor Category", "Item ID", 
            "Description", "New Margin", "Price Levels", "Notes"
        ]
        
        # Setup styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        # Adjust column widths
        col_widths = [15, 20, 20, 15, 30, 12, 15, 30]
        for i, width in enumerate(col_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
            
        # Fill data
        for row_idx, update in enumerate(updates, 2):
            # Format margin as percentage
            if isinstance(update['margin'], (int, float)):
                margin_display = f"{update['margin']:.1%}"
            else:
                margin_display = str(update['margin'])
                
            # Format price levels
            price_levels = ",".join(update['price_levels']) if len(update['price_levels']) < 12 else "ALL"
            
            # Write row
            sheet.cell(row=row_idx, column=1).value = update['zone']
            sheet.cell(row=row_idx, column=2).value = update['category']
            sheet.cell(row=row_idx, column=3).value = update['minor_category']
            sheet.cell(row=row_idx, column=4).value = update['item_id']
            sheet.cell(row=row_idx, column=5).value = update['description']
            sheet.cell(row=row_idx, column=6).value = margin_display
            sheet.cell(row=row_idx, column=7).value = price_levels
            sheet.cell(row=row_idx, column=8).value = update['notes']
            
            # Add light fill to margin cell
            sheet.cell(row=row_idx, column=6).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            # Add borders
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col).border = border
        
        # Add instructions at the top
        instruction_row = sheet.insert_rows(1, 3)
        sheet.merge_cells('A1:H1')
        sheet.merge_cells('A2:H2')
        sheet.merge_cells('A3:H3')
        
        cell = sheet.cell(row=1, column=1)
        cell.value = "RICHMOND BAT FORMAT - COPY AND PASTE TO MARGINS TO CHANGE SHEET"
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        cell = sheet.cell(row=2, column=1)
        cell.value = "Instructions: Select all rows below (including headers) and paste to your Richmond BAT"
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center')
        
        cell = sheet.cell(row=3, column=1)
        cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {len(updates)} items"
        cell.alignment = Alignment(horizontal='center')
    
    def create_lookup_sheet(self, updates):
        """Create lookup table sheet"""
        # Create new sheet
        sheet = self.output_wb.create_sheet(title="Lookup Table")
        
        # Setup headers
        headers = [
            "Pricing Zone", "Product Category", "Minor Category", "Item ID", 
            "Description", "Base Cost", "New Margin", "New Sell Price", "Notes"
        ]
        
        # Setup styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
        # Adjust column widths
        col_widths = [15, 20, 20, 15, 30, 15, 12, 15, 30]
        for i, width in enumerate(col_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
            
        # Add instructions
        sheet.merge_cells('A2:I2')
        cell = sheet.cell(row=2, column=1)
        cell.value = "This sheet is for reference only. Enter base costs to see calculated sell prices."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center')
        
        # Fill data
        for row_idx, update in enumerate(updates, 3):
            # Format margin as percentage
            if isinstance(update['margin'], (int, float)):
                margin_display = f"{update['margin']:.1%}"
            else:
                margin_display = str(update['margin'])
                
            # Write row
            sheet.cell(row=row_idx, column=1).value = update['zone']
            sheet.cell(row=row_idx, column=2).value = update['category']
            sheet.cell(row=row_idx, column=3).value = update['minor_category']
            sheet.cell(row=row_idx, column=4).value = update['item_id']
            sheet.cell(row=row_idx, column=5).value = update['description']
            # Column 6 is for base cost (user can input)
            sheet.cell(row=row_idx, column=7).value = margin_display
            # Column 8 will have a formula
            sheet.cell(row=row_idx, column=9).value = update['notes']
            
            # Add formula for sell price
            base_col = get_column_letter(6)
            margin_cell = sheet.cell(row=row_idx, column=7)
            sell_cell = sheet.cell(row=row_idx, column=8)
            
            # Formula only works with numeric margins
            if isinstance(update['margin'], (int, float)):
                sell_cell.value = f'=IF({base_col}{row_idx}>0,{base_col}{row_idx}/(1-{update["margin"]}),"")'
            else:
                sell_cell.value = ""
                
            # Set formats
            sheet.cell(row=row_idx, column=6).number_format = '$#,##0.00'
            sheet.cell(row=row_idx, column=8).number_format = '$#,##0.00'
            
            # Add light fill to editable cells
            sheet.cell(row=row_idx, column=6).fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Add borders
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row_idx, column=col).border = border
    
    def save_output(self):
        """Save the output file"""
        # Create output name with timestamp
        base_path = os.path.dirname(self.file_path)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = os.path.join(base_path, f"PRICING_OUTPUT_{timestamp}.xlsx")
        
        print(f"\n💾 Saving output file...")
        print(f"   Output: {os.path.basename(output_file)}")
        
        try:
            self.output_wb.save(output_file)
            print(f"✔ File saved successfully")
            
            print("\n" + "="*70)
            print("GENERATION COMPLETE!")
            print("="*70)
            print(f"Input:   {os.path.basename(self.file_path)}")
            print(f"Output:  {os.path.basename(output_file)}")
            print(f"Contains formatting for both Holt and Richmond BAT files")
            print("\nInstructions:")
            print("1. Open the output file")
            print("2. Go to the sheet for your target BAT (Holt or Richmond)")
            print("3. Copy the data including headers")
            print("4. Paste to your target BAT file in the margin updates sheet")
            print("5. Run the normal update process in your BAT file")
            
            # Update status in original file
            if 'Margin Input' in self.wb.sheetnames:
                try:
                    status_sheet = self.wb['Margin Input']
                    # Find a status cell or create one
                    status_cell = status_sheet.cell(row=1, column=11)
                    status_cell.value = f"Output generated: {os.path.basename(output_file)}"
                    status_cell.font = Font(bold=True)
                    
                    # Save the input file with status
                    self.wb.save(self.file_path)
                except:
                    pass
            
            return True
        except Exception as e:
            print(f"❌ Error saving file: {e}")
            return False


def main():
    # Default file path - use local path
    default_path = r"C:\Users\corey.boser\Documents\Holt Monthly Pricing\QUICK_MARGIN_CHANGER.xlsm"
    
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = default_path
        
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        print("\nPlease provide the correct file path:")
        print("python quick_margin_changer.py 'path\\to\\QUICK_MARGIN_CHANGER.xlsm'")
        sys.exit(1)
    
    # Run updater
    changer = QuickMarginChanger(file_path)
    success = changer.run_update()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
